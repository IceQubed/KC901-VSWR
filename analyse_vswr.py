#!/usr/bin/env python3
"""
KC901V VNA S11/VSWR analyser.

Reads .ini and .xlsx measurement files from the input folder, plots VSWR vs
frequency with a highlighted band of interest, and computes a figure of merit
per file (best antenna = VSWR close to 1 over the range of interest).
"""

import argparse
import csv
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator
from openpyxl import load_workbook


def parse_ini(path: Path) -> tuple[np.ndarray, np.ndarray]:
    """Parse a KC901V .ini file; return (freq_Hz, vswr) arrays."""
    freqs = []
    vswr = []
    in_curve = False
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line == "[Curve]":
                in_curve = True
                continue
            if in_curve:
                if line.startswith("["):
                    break
                if line.startswith("F:"):
                    part = line[2:].split(",", 1)
                    if len(part) == 2:
                        freqs.append(int(part[0]))
                        vswr.append(float(part[1]))
    if not freqs:
        raise ValueError(f"No curve data in {path}")
    return np.array(freqs), np.array(vswr)


def parse_xlsx(path: Path) -> tuple[np.ndarray, np.ndarray]:
    """Parse a KC901V .xlsx file (e.g. S11-VSWR sheet); return (freq_Hz, vswr) arrays."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "S11-VSWR" if "S11-VSWR" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    freqs = []
    vswr = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or row[0] is None:
            continue
        # Skip header row (e.g. "Frequency(Hz)", "Value", ...)
        try:
            f = int(float(row[0]))
            v = float(row[1])
        except (TypeError, ValueError, IndexError):
            continue
        freqs.append(f)
        vswr.append(v)
    wb.close()
    if not freqs:
        raise ValueError(f"No curve data in {path}")
    return np.array(freqs), np.array(vswr)


def figure_of_merit(
    freq_Hz: np.ndarray,
    vswr: np.ndarray,
    fmin_Hz: float,
    fmax_Hz: float,
    extra_fmin_Hz: float | None = None,
    extra_fmax_Hz: float | None = None,
    extra_weight: float = 2.0,
) -> dict:
    """
    Compute figures of merit for the band [fmin_Hz, fmax_Hz].
    Best antenna: VSWR close to 1 over the whole range.
    If extra_fmin_Hz/extra_fmax_Hz are set, points in that sub-band are weighted
    by extra_weight in the score (e.g. 2.0 = twice as important).
    Returns dict with mean_vswr, max_vswr, score (0–1, 1 = best).
    """
    mask = (freq_Hz >= fmin_Hz) & (freq_Hz <= fmax_Hz)
    if not np.any(mask):
        return {"mean_vswr": np.nan, "max_vswr": np.nan, "score": 0.0, "n_points": 0}

    f_masked = freq_Hz[mask]
    v = vswr[mask]
    n = len(v)
    mean_vswr = float(np.mean(v))
    max_vswr = float(np.max(v))

    # Weights: 1 in main band; extra_weight in [extra_fmin, extra_fmax] if specified
    use_extra = (
        extra_fmin_Hz is not None
        and extra_fmax_Hz is not None
        and extra_fmin_Hz < extra_fmax_Hz
    )
    if use_extra:
        in_extra = (f_masked >= extra_fmin_Hz) & (f_masked <= extra_fmax_Hz)
        weights = np.where(in_extra, extra_weight, 1.0)
        weighted_sq_err = np.sum(weights * (v - 1.0) ** 2) / np.sum(weights)
        score = 1.0 / (1.0 + weighted_sq_err)
    else:
        mse = np.mean((v - 1.0) ** 2)
        score = 1.0 / (1.0 + mse)

    return {
        "mean_vswr": mean_vswr,
        "max_vswr": max_vswr,
        "score": score,
        "n_points": n,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analyse KC901V VNA S11/VSWR .ini and .xlsx files: plot curves and compute FOM."
    )
    parser.add_argument(
        "--fmin",
        type=float,
        default=1.7e9,
        metavar="Hz",
        help="Minimum frequency of interest in Hz (default: 1.7e9 = 1.7 GHz)",
    )
    parser.add_argument(
        "--fmax",
        type=float,
        default=2.5e9,
        metavar="Hz",
        help="Maximum frequency of interest in Hz (default: 2.5e9 = 2.5 GHz)",
    )
    parser.add_argument(
        "--extra-fmin",
        type=float,
        default=None,
        metavar="Hz",
        help="Start of extra-scoring sub-band in Hz (must be within --fmin/--fmax)",
    )
    parser.add_argument(
        "--extra-fmax",
        type=float,
        default=None,
        metavar="Hz",
        help="End of extra-scoring sub-band in Hz (must be within --fmin/--fmax)",
    )
    parser.add_argument(
        "--extra-weight",
        type=float,
        default=2.0,
        metavar="W",
        help="Weight for points in extra-scoring sub-band (default: 2 = twice as important)",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("input"),
        help="Folder containing .ini and/or .xlsx measurement files (default: input)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output"),
        help="Folder for all outputs: plot and FOM report (default: output)",
    )
    parser.add_argument(
        "--plot-file",
        type=str,
        default="vswr_curves.png",
        help="Plot filename inside output folder (default: vswr_curves.png)",
    )
    args = parser.parse_args()

    fmin_hz = args.fmin
    fmax_hz = args.fmax
    if fmin_hz >= fmax_hz:
        parser.error("--fmin must be less than --fmax")

    # Extra-scoring is only used when user explicitly passes both --extra-fmin and --extra-fmax
    extra_fmin = args.extra_fmin
    extra_fmax = args.extra_fmax
    extra_weight = args.extra_weight
    if (extra_fmin is not None) != (extra_fmax is not None):
        parser.error("--extra-fmin and --extra-fmax must be given together")
    if extra_fmin is not None and extra_fmax is not None:
        if extra_fmin >= extra_fmax:
            parser.error("--extra-fmin must be less than --extra-fmax")
        if extra_fmin < fmin_hz or extra_fmax > fmax_hz:
            parser.error("Extra-scoring band must lie within region of interest (--fmin to --fmax)")
        if extra_weight <= 0:
            parser.error("--extra-weight must be positive")

    input_dir = args.input
    output_dir = args.output
    if not input_dir.is_dir():
        parser.error(f"Input folder not found: {input_dir}")

    data_files = sorted(input_dir.glob("*.ini")) + sorted(input_dir.glob("*.xlsx"))
    if not data_files:
        print(f"No .ini or .xlsx files found in {input_dir}")
        return

    # Load all curves and FOMs (same handling for .ini and .xlsx)
    all_data: list[tuple[str, np.ndarray, np.ndarray, dict]] = []
    for p in data_files:
        try:
            if p.suffix.lower() == ".ini":
                freq, vswr = parse_ini(p)
            else:
                freq, vswr = parse_xlsx(p)
            if extra_fmin is not None and extra_fmax is not None:
                fom = figure_of_merit(
                    freq, vswr, fmin_hz, fmax_hz,
                    extra_fmin_Hz=extra_fmin,
                    extra_fmax_Hz=extra_fmax,
                    extra_weight=extra_weight,
                )
            else:
                fom = figure_of_merit(freq, vswr, fmin_hz, fmax_hz)
            all_data.append((p.stem, freq, vswr, fom))
        except Exception as e:
            print(f"Warning: skip {p.name}: {e}")

    if not all_data:
        print("No valid measurement data loaded.")
        return

    # Order best to worst (by score descending)
    by_score = sorted(all_data, key=lambda x: x[3]["score"], reverse=True)

    # Plot (same order: best to worst in legend)
    output_dir.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(12, 12))

    n_series = max(len(by_score), 1)
    colors = plt.cm.nipy_spectral(np.linspace(0.1, 0.95, n_series))
    for i, (name, freq, vswr, fom) in enumerate(by_score):
        freq_ghz = freq / 1e9
        label = f"{name} (FoM {fom['score']:.3f})"
        ax.plot(freq_ghz, vswr, label=label, color=colors[i % len(colors)], alpha=0.9)

    # Highlight band of interest
    ax.axvspan(fmin_hz / 1e9, fmax_hz / 1e9, alpha=0.15, color="green", zorder=0)
    ax.axvline(fmin_hz / 1e9, color="green", linestyle="--", alpha=0.6, linewidth=0.8)
    ax.axvline(fmax_hz / 1e9, color="green", linestyle="--", alpha=0.6, linewidth=0.8)
    # Extra-scoring sub-band (darker overlay)
    if extra_fmin is not None and extra_fmax is not None:
        ax.axvspan(
            extra_fmin / 1e9, extra_fmax / 1e9,
            alpha=0.25, color="green", zorder=0,
        )
        ax.axvline(extra_fmin / 1e9, color="green", linestyle=":", alpha=0.7, linewidth=0.8)
        ax.axvline(extra_fmax / 1e9, color="green", linestyle=":", alpha=0.7, linewidth=0.8)

    ax.set_xlabel("Frequency (GHz)")
    ax.set_ylabel("VSWR")
    title = "VSWR vs frequency (band of interest shaded)"
    if extra_fmin is not None and extra_fmax is not None:
        title += f"\n(darker: extra-scoring {extra_fmin/1e9:.2f}–{extra_fmax/1e9:.2f} GHz, weight {extra_weight}×)"
    ax.set_title(title)
    ax.xaxis.set_major_locator(MultipleLocator(0.1))
    ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1), fontsize=8, frameon=True)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(bottom=0.95)
    fig.tight_layout(rect=(0, 0, 0.92, 1))
    # Save plot to output folder (bbox_inches='tight' crops unused whitespace)
    plot_path = output_dir / args.plot_file
    fig.savefig(plot_path, dpi=150, bbox_inches="tight", pad_inches=0.2)
    plt.close(fig)

    # Build and save figure-of-merit report to output folder (best to worst)
    fom_path = output_dir / "vswr_fom.txt"
    with open(fom_path, "w", encoding="utf-8") as report:
        report.write(
            "Figure of merit (band of interest: {:.3f}–{:.3f} GHz)\n".format(
                fmin_hz / 1e9, fmax_hz / 1e9
            )
        )
        if extra_fmin is not None and extra_fmax is not None:
            report.write(
                "Extra-scoring sub-band: {:.3f}–{:.3f} GHz (weight {}×)\n".format(
                    extra_fmin / 1e9, extra_fmax / 1e9, extra_weight
                )
            )
        report.write("-" * 70 + "\n")
        report.write(f"{'File':<45} {'Mean VSWR':>10} {'Max VSWR':>10} {'Score':>8}\n")
        report.write("-" * 70 + "\n")
        for name, _freq, _vswr, fom in by_score:
            report.write(
                f"{name:<45} {fom['mean_vswr']:>10.4f} {fom['max_vswr']:>10.4f} {fom['score']:>8.4f}\n"
            )
        report.write("-" * 70 + "\n")
        if extra_fmin is not None and extra_fmax is not None:
            report.write(
                "Score: weighted by sub-band (extra band {}×). 1 = ideal. Lower mean/max VSWR is better.\n".format(extra_weight)
            )
        else:
            report.write(
                "Score: 1 = ideal (VSWR = 1 across band). Lower mean/max VSWR is better.\n"
            )
        best = by_score[0]
        report.write(f"Best in band: {best[0]} (score {best[3]['score']:.4f})\n")

    # Save same figure-of-merit data as CSV
    fom_csv_path = output_dir / "vswr_fom.csv"
    with open(fom_csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["File", "Mean VSWR", "Max VSWR", "Score"])
        for name, _freq, _vswr, fom in by_score:
            writer.writerow([name, f"{fom['mean_vswr']:.4f}", f"{fom['max_vswr']:.4f}", f"{fom['score']:.4f}"])

    # Print summary to terminal and confirm all outputs are in output folder
    print("\nFigure of merit (band of interest: {:.3f}–{:.3f} GHz)".format(fmin_hz / 1e9, fmax_hz / 1e9))
    if extra_fmin is not None and extra_fmax is not None:
        print("Extra-scoring sub-band: {:.3f}–{:.3f} GHz (weight {}×)".format(extra_fmin / 1e9, extra_fmax / 1e9, extra_weight))
    print("-" * 70)
    print(f"{'File':<45} {'Mean VSWR':>10} {'Max VSWR':>10} {'Score':>8}")
    print("-" * 70)
    for name, _freq, _vswr, fom in by_score:
        print(
            f"{name:<45} {fom['mean_vswr']:>10.4f} {fom['max_vswr']:>10.4f} {fom['score']:>8.4f}"
        )
    print("-" * 70)
    if extra_fmin is not None and extra_fmax is not None:
        print("Score: weighted by sub-band (extra band {}×). 1 = ideal. Lower mean/max VSWR is better.".format(extra_weight))
    else:
        print("Score: 1 = ideal (VSWR = 1 across band). Lower mean/max VSWR is better.")
    print(f"Best in band: {best[0]} (score {best[3]['score']:.4f})")
    print(f"\nAll outputs written to {output_dir.resolve()}:")
    print(f"  - {plot_path.name}")
    print(f"  - {fom_path.name}")
    print(f"  - {fom_csv_path.name}")


if __name__ == "__main__":
    main()
